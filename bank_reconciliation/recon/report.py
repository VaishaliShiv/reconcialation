"""Shared result-dataset builder + writers for the NEW-format pipeline.

Both the CSV runner (scripts/recon_sample.py) and the Cosmos runner
(scripts/recon_cosmos.py) feed canonical txns in here so the reconcile → build →
write logic lives in exactly one place. No LLM — deterministic money math.
"""
from __future__ import annotations

import csv
from datetime import datetime, timezone

from ..models import MatchType
from ..schema import canonical_mapper as cmap
from . import actions, balance
from . import dynamic_matcher as matcher

COLUMNS = [
    "run_id", "bank_name", "source_channel", "settlement_date", "match_key_used",
    "partner_trn_reference_no", "payment_ref_no", "dewa_trn_reference_no", "type",
    "amount_source", "amount_sap", "amount_diff", "date_source", "date_sap",
    "classification", "action", "comment", "status", "recon_status",
]
NUMERIC = {"amount_source", "amount_sap", "amount_diff"}


def run_id_for(vendor: str) -> str:
    return f"{vendor}-{datetime.now(timezone.utc):%Y%m%dT%H%M%S}"


def _cell(v) -> str:
    return "" if v is None else str(v)


def reconcile_and_build(file_txns, sap_txns, vendor, run_id, one_to_many=False):
    """Return (rows, snapshot, meta). `completed` file rows are skipped (not re-queried)."""
    active = [t for t in file_txns if (t.status or "").lower() != "completed"]
    it_date = sorted({t.txn_date.isoformat() for t in active if t.txn_date})
    raw = matcher.reconcile(active, sap_txns, one_to_many=one_to_many)
    snap = balance.snapshot(vendor, [t for t in active if t.valid], sap_txns, raw)
    settlement = next((t.settlement_date.isoformat() for t in active if t.settlement_date), "")

    rows: list[dict] = []
    for r in raw:
        act, comment = actions.decide(r["match_type"])
        if r.get("note"):
            comment = f"{comment} [{r['note']}]"
        matched = r["match_type"] == MatchType.MATCHED
        rows.append({
            "run_id": run_id, "bank_name": vendor,
            "source_channel": r.get("source_channel") or cmap.VENDOR_SOURCE.get(vendor, vendor),
            "settlement_date": _cell(r.get("settlement_date")) or settlement,
            "match_key_used": r.get("match_key_used"),
            "partner_trn_reference_no": r.get("partner_trn_reference_no"),
            "payment_ref_no": r.get("payment_ref_no"),
            "dewa_trn_reference_no": r.get("dewa_trn_reference_no"),
            "type": r.get("type"),
            "amount_source": r.get("file_amount"), "amount_sap": r.get("sap_amount"),
            "amount_diff": r.get("amount_diff"),
            "date_source": _cell(r.get("date_source")), "date_sap": _cell(r.get("date_sap")),
            "classification": r["match_type"].value, "action": act.value, "comment": comment,
            "status": "completed" if matched else "anomaly",
            "recon_status": snap.recon_status.value,
        })
    meta = {"file": len(file_txns), "active": len(active),
            "skipped": len(file_txns) - len(active), "sap": len(sap_txns),
            "it_date": it_date, "raw": raw}
    return rows, snap, meta


def sap_row_statuses(raw, sap_txns) -> list[str]:
    """Per-SAP-row status aligned to `sap_txns` order: 'completed' if that row's
    join key matched cleanly, else 'anomaly' (mismatch / duplicate / missing_in_file)."""
    key_class: dict[str, MatchType] = {}
    for r in raw:
        ku = r.get("match_key_used") or ""
        val = ku.split(":", 1)[1] if ":" in ku else ku
        if val:
            key_class[val] = r["match_type"]
    return ["completed" if key_class.get(s.partner_txn_id) == MatchType.MATCHED
            else "anomaly" for s in sap_txns]


def write_sap_status(path, statuses) -> None:
    """Rewrite the SAP CSV in place, filling only the `status` column (order preserved).

    All other cells are copied through verbatim so amounts/formatting are untouched.
    """
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        fields = list(reader.fieldnames or [])
        rows = list(reader)
    if "status" not in fields:
        fields.append("status")
    for row, st in zip(rows, statuses):
        row["status"] = st
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def print_summary(rows, snap, meta, vendor):
    print(f"file rows={meta['file']}  active={meta['active']}  "
          f"skipped(completed)={meta['skipped']}  sap rows={meta['sap']}\n")
    print(f"=== Reconciliation Result: {vendor} ({len(rows)} rows) ===")
    print(f"{'key_used':<24}{'src':>10}{'sap':>10}{'diff':>9}  {'classification':<16}{'action'}")
    for r in rows:
        print(f"{_cell(r['match_key_used']):<24}{_cell(r['amount_source']):>10}"
              f"{_cell(r['amount_sap']):>10}{_cell(r['amount_diff']):>9}  "
              f"{r['classification']:<16}{r['action']}")
    by: dict[str, int] = {}
    for r in rows:
        by[r["classification"]] = by.get(r["classification"], 0) + 1
    print(f"\nbreakdown : {by}")
    print(f"balance   : file {snap.file_total} | sap {snap.sap_total} | diff {snap.balance_diff}")
    print(f"status    : {snap.recon_status.value}")


def write_csv(rows, path, columns=None) -> None:
    columns = columns or COLUMNS
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows, path, columns=None) -> None:
    """Formatted Excel. ID columns are TEXT so big numbers don't become 2.6E+18."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = columns or COLUMNS
    wb = Workbook()
    ws = wb.active
    ws.title = "reconciliation"
    head_font, head_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="0F4024")
    anomaly_fill = PatternFill("solid", fgColor="FCE8E6")
    for c, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font, cell.fill, cell.alignment = head_font, head_fill, Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        is_anom = row.get("classification") != "matched"
        for c, col in enumerate(columns, 1):
            v = row.get(col)
            cell = ws.cell(row=r, column=c)
            if col in NUMERIC and v is not None and v != "":
                cell.value, cell.number_format = float(v), "#,##0.00"
            else:
                cell.value, cell.number_format = _cell(v), "@"
            if is_anom:
                cell.fill = anomaly_fill
    for c, col in enumerate(columns, 1):
        widest = max([len(col)] + [len(_cell(row.get(col))) for row in rows])
        ws.column_dimensions[get_column_letter(c)].width = min(widest + 2, 45)
    ws.freeze_panes = "A2"
    wb.save(path)
